import requests
import re
import openpyxl
import os
import html


def remove_html_tags(info_dict):
    """去除字典值中的HTML标签"""
    pattern = re.compile(r"<[^>]+>")
    return {k: pattern.sub("", v) for k, v in info_dict.items()}


def remove_pinyin(info_dict):
    """去掉中括号的拼音"""
    pattern = re.compile(r"\[.*?\]")
    return {k: pattern.sub("", v) for k, v in info_dict.items()}


# 恢复HTML转义
def restore_html_escape(info_dict):
    """恢复HTML转义"""
    return {k: html.unescape(v) for k, v in info_dict.items()}


# 初始化Excel
def init_excel():
    """初始化Excel，创建多个名为'常用中药'、'经典方剂'、'中成药'和'药膳'的Excel文件"""

    # 定义要创建的文件和对应的Sheet名称
    files_and_sheets = {
        "常用中药.xlsx": "常用中药",
        "经典方剂.xlsx": "经典方剂",
        "中成药.xlsx": "中成药",
        "药膳.xlsx": "药膳",
    }

    for filename, sheet_name in files_and_sheets.items():
        # 检测是否存在文件，存在则跳过
        if os.path.exists(filename):
            print(f"{filename}文件已存在，跳过初始化")
            continue
        # 创建一个新的Excel文件
        workbook = openpyxl.Workbook()
        # 创建一个新的Sheet
        workbook.create_sheet(sheet_name, 0)
        # 保存Excel文件
        workbook.save(filename)
        print(f"{filename}文件初始化完成")


# 添加内容到常用中药Excel
def add_content_to_common_medicine_excel(info_dict):
    """根据字典内容添加到Excel"""
    workbook = openpyxl.load_workbook("常用中药.xlsx")
    sheet = workbook["常用中药"]  # 直接通过sheet名称获取

    # 如果是第一次写入，添加表头
    if sheet.max_row == 1:
        sheet.append(list(info_dict.keys()))

    # 添加字典的值
    sheet.append(list(info_dict.values()))
    # 保存Excel文件
    workbook.save("常用中药.xlsx")
    print(f"添加药品信息到Excel完成")
    print("-" * 50)


# 添加内容到经典方剂Excel
def add_content_to_prescription_excel(info_dict):
    """根据字典内容添加到Excel"""
    workbook = openpyxl.load_workbook("经典方剂.xlsx")
    sheet = workbook["经典方剂"]  # 直接通过sheet名称获取

    # 如果是第一次写入，添加表头
    if sheet.max_row == 1:
        sheet.append(list(info_dict.keys()))

    # 添加字典的值
    sheet.append(list(info_dict.values()))
    # 保存Excel文件
    workbook.save("经典方剂.xlsx")
    print(f"添加经典方剂信息到Excel完成")
    print("-" * 50)


# 添加内容到中成药Excel
def add_content_to_medicine_excel(info_dict):
    """根据字典内容添加到Excel"""
    workbook = openpyxl.load_workbook("中成药.xlsx")
    sheet = workbook["中成药"]  # 直接通过sheet名称获取

    # 如果是第一次写入，添加表头
    if sheet.max_row == 1:
        sheet.append(list(info_dict.keys()))

    # 添加字典的值
    sheet.append(list(info_dict.values()))
    # 保存Excel文件
    workbook.save("中成药.xlsx")
    print(f"添加中成药信息到Excel完成")
    print("-" * 50)


# 添加内容到药膳Excel
def add_content_to_diet_excel(info_dict):
    """根据字典内容添加到Excel"""
    workbook = openpyxl.load_workbook("药膳.xlsx")
    sheet = workbook["药膳"]  # 直接通过sheet名称获取

    # 如果是第一次写入，添加表头
    if sheet.max_row == 1:
        sheet.append(list(info_dict.keys()))

    # 添加字典的值
    sheet.append(list(info_dict.values()))
    # 保存Excel文件
    workbook.save("药膳.xlsx")
    print(f"添加药膳信息到Excel完成")
    print("-" * 50)


def fetch_webpage(url):
    """获取网页内容并设置编码"""
    response = requests.get(url)
    response.encoding = "utf-8"
    return response.text


def save_response(content, filename="response.html"):
    """保存响应内容到文件"""
    with open(filename, "w", encoding="utf-8") as f:
        f.write(content)


# 提取常用中药ID和名称
def extract_common_medicine_info(text_content):
    """提取常用中药ID和名称"""
    pattern = re.compile(
        r'<a href="/traditionaldetails\?id=(\d+)" target="_blank" data-v-bce4c468>(.*?)</a>'
    )
    return pattern.findall(text_content.replace("\n", ""))


# 提取经典方剂ID和名称
def extract_prescription_info(text_content):
    """提取经典方剂ID和名称"""
    pattern = re.compile(
        r'<a href="/prescriptiondetails\?id=(\d+)" target="_blank" data-v-612dcc6e>(.*?)</a>'
    )
    return pattern.findall(text_content.replace("\n", ""))


# 提取中成药ID和名称
def extract_chinese_medicine_info(text_content):
    """提取中成药ID和名称"""
    pattern = re.compile(
        r'<div title="(.*?)" class="pharmacy_item"[^>]*><a href="/chinesemedicinedetails\?id=(\d+)"[^>]*>.*?<div class="in_box"[^>]*>.*?</div>'
    )
    return pattern.findall(text_content.replace("\n", ""))


# 提取药膳ID和名称
def extract_diet_info(text_content):
    """提取药膳ID和名称"""
    pattern = re.compile(
        r'<div title=".*?" class="prescriptionsLists_item"[^>]*><a href="/tonicdietdetails\?id=(\d+)"[^>]*><div class="out_box"[^>]*><div class="in_box"[^>]*>.*?<img[^>]*>(.*?)<img'
    )
    return pattern.findall(text_content.replace("\n", ""))


# 处理常用中药和经典方剂和药膳的ID和名称
def process_common_medicine_and_prescription_and_diet_id_name(matches):
    """处理常用中药和经典方剂和药膳的ID和名称"""
    ids = [match[0].strip() for match in matches]
    medicine_names = [match[1].strip() for match in matches]
    return ids, medicine_names


# 处理中成药的ID和名称
def process_chinese_medicine_id_name(matches):
    """处理中成药的ID和名称"""
    medicine_names = [match[0].strip() for match in matches]
    ids = [match[1].strip() for match in matches]
    return ids, medicine_names


def print_medicine_info(matches):
    """打印药品信息"""
    for id_num, medicine_name in matches:
        print(f"ID: {id_num}, 药名: {medicine_name.strip()}")


# 获取常用中药ID和名称映射
def get_common_medicine_id_map():
    """获取常用中药ID和名称映射"""
    url = "https://www.zhiyuanzhongyi.com/traditional"
    # 获取内容
    content = fetch_webpage(url)
    # 提取药品信息
    matches = extract_common_medicine_info(content)
    # 获取分离的ID和药名列表
    ids, medicine_names = process_common_medicine_and_prescription_and_diet_id_name(
        matches
    )
    print(f"获取常用中药ID和名称映射完成，共获取到{len(ids)}个")
    return ids, medicine_names


# 获取经典方剂ID和名称映射
def get_prescription_id_map():
    url = "https://www.zhiyuanzhongyi.com/prescription"
    # 获取内容
    content = fetch_webpage(url)
    # 提取药品信息
    matches = extract_prescription_info(content)
    # 获取分离的ID和药名列表
    ids, medicine_names = process_common_medicine_and_prescription_and_diet_id_name(
        matches
    )
    print(f"获取经典方剂ID和名称映射完成，共获取到{len(ids)}个经典方剂")
    return ids, medicine_names


# 获取中成药ID和名称映射
def get_chinese_medicine_id_map():
    """获取中成药ID和名称映射"""
    url = "https://www.zhiyuanzhongyi.com/pharmacy"
    content = fetch_webpage(url)
    matches = extract_chinese_medicine_info(content)
    ids, medicine_names = process_chinese_medicine_id_name(matches)
    print(f"获取中成药ID和名称映射完成，共获取到{len(ids)}个中成药")
    return ids, medicine_names


# 获取药膳ID和名称映射
def get_diet_id_map():
    """获取药膳ID和名称映射"""
    url = "https://www.zhiyuanzhongyi.com/tonicdiet"
    content = fetch_webpage(url)
    matches = extract_diet_info(content)
    ids, medicine_names = process_common_medicine_and_prescription_and_diet_id_name(
        matches
    )
    print(f"获取药膳ID和名称映射完成，共获取到{len(ids)}个药膳")
    return ids, medicine_names


# 处理常用中药信息
def process_common_medicine_info(ids, medicine_names):
    """处理常用中药信息"""
    base_url = "https://www.zhiyuanzhongyi.com/traditionaldetails?id="
    print("开始处理常用中药信息")
    total = len(ids)  # 获取总数
    for index, (id, medicine_name) in enumerate(zip(ids, medicine_names), start=1):
        url = base_url + id
        print(
            f"({index}/{total} ({(index/total)*100:.2f}%))开始处理常用中药信息：{medicine_name}，id：{id}"
        )
        response = fetch_webpage(url)

        # 去除换行符
        response = response.replace("\n", "")

        # 提取药品所有属性名称存成列表
        pattern_info_name = re.compile(
            r'<div class="left_title" data-v-0bab2978>(.*?)</div>'
        )

        # 提取属性对应的数据存成列表
        pattern_info_data = re.compile(
            r'<div class="right_msg" data-v-0bab2978>(.*?)</div>'
        )

        matches_info_name = pattern_info_name.findall(response)
        matches_info_data = pattern_info_data.findall(response)
        if matches_info_name and matches_info_data:
            # 拼成字典
            info_dict = dict(zip(matches_info_name, matches_info_data))
            print(f"提取常用中药信息：{medicine_name}，id：{id}完成")
            # 去掉HTML标签
            info_dict = remove_html_tags(info_dict)
            print(f"去除HTML标签完成")
            # 去掉中括号的拼音
            info_dict = remove_pinyin(info_dict)
            print(f"去掉中括号的拼音完成")
            # 添加到Excel
            add_content_to_common_medicine_excel(info_dict)


# 处理经典方剂信息
def process_prescription_info(ids, medicine_names):
    base_url = "https://www.zhiyuanzhongyi.com/prescriptiondetails?id="
    print("开始处理经典方剂信息")
    total = len(ids)  # 获取总数
    for index, (id, medicine_name) in enumerate(zip(ids, medicine_names), start=1):
        url = base_url + id
        print(
            f"({index}/{total} ({(index/total)*100:.2f}%))开始处理经典方剂信息：{medicine_name}，id：{id}"
        )
        response = fetch_webpage(url)

        # 去除换行符
        response = response.replace("\n", "")

        # 提取经典方剂所有属性名称存成列表
        pattern_info_name = re.compile(
            r'<div class="left_title" data-v-470d2d5c>(.*?)</div>'
        )

        # 提取属性对应的数据存成列表
        pattern_info_data = re.compile(
            r'<div class="right_msg" data-v-470d2d5c>(.*?)</div>'
        )

        matches_info_name = pattern_info_name.findall(response)
        matches_info_data = pattern_info_data.findall(response)
        if matches_info_name and matches_info_data:
            # 拼成字典
            info_dict = dict(zip(matches_info_name, matches_info_data))
            print(f"提取经典方剂信息：{medicine_name}，id：{id}完成")
            # 去掉HTML标签
            info_dict = remove_html_tags(info_dict)
            print(f"去除HTML标签完成")
            # 去掉中括号的拼音
            info_dict = remove_pinyin(info_dict)
            print(f"去掉中括号的拼音完成")
            # 添加到Excel
            add_content_to_prescription_excel(info_dict)


# 处理中成药信息
def process_chinese_medicine_info(ids, medicine_names):
    """处理中成药信息"""
    base_url = "https://www.zhiyuanzhongyi.com/chinesemedicinedetails?id="
    print("开始处理中成药信息")
    total = len(ids)  # 获取总数
    for index, (id, medicine_name) in enumerate(zip(ids, medicine_names), start=1):
        url = base_url + id
        print(
            f"({index}/{total} ({(index/total)*100:.2f}%))开始处理中成药信息：{medicine_name}，id：{id}"
        )
        response = fetch_webpage(url)

        # 去除换行符
        response = response.replace("\n", "")

        # 匹配中成药名字
        pattern_medicine_name = re.compile(
            r'<div class="title_msg" data-v-e8e5f53c>(.*?)</div>'
        )
        matches_medicine_name = pattern_medicine_name.findall(response)

        if matches_medicine_name:
            medicine_name = matches_medicine_name[0].strip()

        # 提取中成药数据键
        pattern_info_key = re.compile(
            r'<div class="left_title" data-v-a7c42b58>(.*?)</div>'
        )

        # 提取中成药数据值
        pattern_info_data = re.compile(
            r'<div class="right_msg" data-v-a7c42b58>(.*?)</div>'
        )

        matches_info_key = pattern_info_key.findall(response)
        matches_info_data = pattern_info_data.findall(response)
        if matches_info_key and matches_info_data:
            # 把中成药名字添加到字典中
            info_dict = {"药品名称": medicine_name}
            # 拼成字典
            info_dict.update(dict(zip(matches_info_key, matches_info_data)))
            print(f"提取中成药信息：{medicine_name}，id：{id}完成")
            # 去掉HTML标签
            info_dict = remove_html_tags(info_dict)
            print(f"去除HTML标签完成")
            # 去掉中括号的拼音
            info_dict = remove_pinyin(info_dict)
            print(f"去掉中括号的拼音完成")
            # 添加到Excel
            add_content_to_medicine_excel(info_dict)


# 处理药膳信息
def process_diet_info(ids, medicine_names):
    """处理药膳信息"""
    base_url = "https://www.zhiyuanzhongyi.com/tonicdietdetails?id="
    print("开始处理药膳信息")
    total = len(ids)
    for index, (id, medicine_name) in enumerate(zip(ids, medicine_names), start=1):
        url = base_url + id
        print(
            f"({index}/{total} ({(index/total)*100:.2f}%))开始处理药膳信息：{medicine_name} ，id：{id}"
        )
        response = fetch_webpage(url)
        response = response.replace("\n", "")

        # 匹配药膳名字
        pattern_medicine_name = re.compile(
            r'<div class="title_msg" data-v-e8e5f53c>(.*?)</div>'
        )

        matches_medicine_name = pattern_medicine_name.findall(response)

        if matches_medicine_name:
            medicine_name = matches_medicine_name[0].strip()

        # 数据键
        pattern_info_key = re.compile(
            r'<div class="left_title" data-v-a7c42b58>(.*?)</div>'
        )

        # 数据值
        pattern_info_data = re.compile(
            r'<div class="right_msg" data-v-a7c42b58>(.*?)</div>'
        )
        matches_info_key = pattern_info_key.findall(response)
        matches_info_data = pattern_info_data.findall(response)

        if matches_info_key and matches_info_data:
            # 把药膳名字添加到字典中
            info_dict = {"药品名称": medicine_name}
            # 拼成字典
            info_dict.update(dict(zip(matches_info_key, matches_info_data)))
            print(f"提取药膳信息：{medicine_name}，id：{id}完成")
            # 去掉HTML标签
            info_dict = remove_html_tags(info_dict)
            print(f"去除HTML标签完成")
            # 去掉中括号的拼音
            info_dict = remove_pinyin(info_dict)
            print(f"去掉中括号的拼音完成")
            add_content_to_diet_excel(info_dict)


def process_common_medicine():
    """处理常用中药"""
    print("-" * 20 + "常用中药" + "-" * 20 + "by W1ndys")
    ids, medicine_names = get_common_medicine_id_map()
    print("-" * 50)
    process_common_medicine_info(ids, medicine_names)
    print("-" * 50)


def process_prescription():
    """处理经典方剂"""
    print("-" * 20 + "经典方剂" + "-" * 20 + "by W1ndys")
    ids, medicine_names = get_prescription_id_map()
    print("-" * 50)
    process_prescription_info(ids, medicine_names)
    print("-" * 50)


def process_chinese_medicine():
    """处理中成药"""
    print("-" * 20 + "中成药" + "-" * 20 + "by W1ndys")
    ids, medicine_names = get_chinese_medicine_id_map()
    print("-" * 50)
    process_chinese_medicine_info(ids, medicine_names)
    print("-" * 50)


def process_diet():
    """处理药膳"""
    print("-" * 20 + "药膳" + "-" * 20 + "by W1ndys")
    ids, medicine_names = get_diet_id_map()
    print("-" * 50)
    process_diet_info(ids, medicine_names)
    print("-" * 50)


if __name__ == "__main__":
    print("-" * 50)
    print("开始执行，Writed by W1ndys，https://github.com/W1ndys/Python_robots_qimo")
    print("-" * 50)
    init_excel()
    print("-" * 50)

    process_common_medicine()
    process_prescription()
    process_chinese_medicine()
    process_diet()
